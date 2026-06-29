from __future__ import annotations

import json
from pathlib import Path
from typing import Any, Mapping

from openpyxl import load_workbook


def _normalize_platform(platform: str) -> str:
    value = str(platform).strip().lower()
    aliases = {
        "boiler": "boiler_ccs",
        "boiler_ccs": "boiler_ccs",
        "te": "tennessee_eastman",
        "tennessee_eastman": "tennessee_eastman",
    }
    try:
        return aliases[value]
    except KeyError as exc:
        raise ValueError(f"unsupported platform: {platform}") from exc


def _platform_dir(root: Path, platform: str) -> Path:
    return Path(root) / "simulators" / _normalize_platform(platform)


def _load_manifest(path: Path) -> dict[str, Any]:
    return json.loads(path.read_text(encoding="utf-8"))


def _load_variable_rows(path: Path) -> list[dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            return []
        headers = [str(value).strip() if value is not None else "" for value in rows[0]]
        records: list[dict[str, str]] = []
        for values in rows[1:]:
            record = {
                headers[index]: str(value).strip() if value is not None else ""
                for index, value in enumerate(values)
                if index < len(headers) and headers[index]
            }
            if any(record.values()):
                records.append(record)
        return records
    finally:
        workbook.close()


def _record_names(record: Mapping[str, Any]) -> set[str]:
    names = {
        str(record.get("canonical_name", "")).strip(),
        str(record.get("manifest_name", "")).strip(),
        str(record.get("runtime_field", "")).strip(),
    }
    names.update(str(alias).strip() for alias in record.get("aliases", []) if str(alias).strip())
    return {name for name in names if name}


def _declared_names(items: list[Any]) -> set[str]:
    names: set[str] = set()
    for item in items:
        if not isinstance(item, Mapping):
            value = str(item).strip()
            if value:
                names.add(value)
            continue
        for key in ("canonical_name", "name", "field", "runtime_field", "manifest_name", "var"):
            value = str(item.get(key, "")).strip()
            if value:
                names.add(value)
    return names


def _normalize_roles(
    record: Mapping[str, Any],
    *,
    injectable_names: set[str],
    observable_names: set[str],
) -> list[str]:
    record_names = _record_names(record)
    roles: list[str] = []
    if record_names & injectable_names:
        roles.append("injectable")
    if record_names & observable_names:
        roles.append("observable")
    return roles


def _canonical_record(
    record: Mapping[str, Any],
    *,
    platform: str,
    roles: list[str],
) -> dict[str, Any]:
    canonical_name = str(record.get("canonical_name", "")).strip()
    manifest_name = str(record.get("manifest_name", "")).strip()
    runtime_field = str(record.get("runtime_field", "")).strip()
    layer = str(record.get("layer", "")).strip().lower()
    controller = str(record.get("controller", "")).strip()
    description = str(record.get("description", "")).strip()
    aliases: list[str] = []
    aliases.extend(
        alias.strip()
        for alias in str(record.get("aliases", "")).split("|")
        if alias.strip() and alias.strip() != canonical_name
    )
    for alias in (manifest_name, runtime_field):
        if alias and alias != canonical_name and alias not in aliases:
            aliases.append(alias)
    return {
        "canonical_name": canonical_name,
        "manifest_name": manifest_name,
        "runtime_field": runtime_field,
        "aliases": aliases,
        "layer": layer,
        "controller": controller,
        "description": description,
        "roles": roles,
        "platform": platform,
    }


def _validate_unique_record_identifiers(records: list[Mapping[str, Any]]) -> None:
    owners: dict[str, str] = {}
    for record in records:
        canonical_name = str(record.get("canonical_name", "")).strip()
        for identifier in _record_names(record):
            owner = owners.get(identifier)
            if owner is not None and owner != canonical_name:
                raise ValueError(
                    f"ambiguous variable identifier {identifier}: {owner}, {canonical_name}"
                )
            owners[identifier] = canonical_name


def _manifest_variable_items(manifest: Mapping[str, Any]) -> list[Mapping[str, Any]]:
    items = [item for item in manifest.get("injection_points", []) if isinstance(item, Mapping)]
    process = manifest.get("physical_process", {})
    if isinstance(process, Mapping):
        items.extend(
            item
            for item in process.get("state_variables", [])
            if isinstance(item, Mapping)
        )
    return items


def _validate_manifest_field_alignment(
    manifest: Mapping[str, Any],
    records: list[Mapping[str, Any]],
) -> None:
    for item in _manifest_variable_items(manifest):
        field = str(item.get("field", "")).strip()
        if not field:
            continue
        item_names = _declared_names([item]) - {field}
        if not item_names:
            continue
        for record in records:
            if not (_record_names(record) & item_names):
                continue
            runtime_field = str(record.get("runtime_field", "")).strip()
            if runtime_field and runtime_field != field:
                name = str(item.get("name", "")).strip() or sorted(item_names)[0]
                raise ValueError(
                    f"manifest field mismatch for {name}: {field} != {runtime_field}"
                )


def _build_boundary_sets(manifest: Mapping[str, Any]) -> tuple[set[str], set[str], set[str]]:
    injection_points = manifest.get("injection_points", [])
    process = manifest.get("physical_process", {})
    state_variables = process.get("state_variables", []) if isinstance(process, Mapping) else []
    injectable_names = _declared_names(list(injection_points))
    observable_names = _declared_names(list(state_variables))
    boundary_names = injectable_names | observable_names
    return boundary_names, injectable_names, observable_names


def build_platform_variable_map(root: Path, platform: str) -> list[dict[str, Any]]:
    platform_name = _normalize_platform(platform)
    platform_dir = _platform_dir(root, platform_name)
    manifest = _load_manifest(platform_dir / "system_manifest.json")
    variable_rows = _load_variable_rows(platform_dir / "variables-table.xlsx")
    boundary_names, injectable_names, observable_names = _build_boundary_sets(manifest)
    records: list[dict[str, Any]] = []
    seen: set[str] = set()
    for row in variable_rows:
        canonical_name = str(row.get("canonical_name", "")).strip()
        if not canonical_name:
            raise ValueError("variable table row missing canonical_name")
        if canonical_name in seen:
            raise ValueError(f"duplicate canonical_name: {canonical_name}")
        if not (_record_names(row) & boundary_names):
            continue
        roles = _normalize_roles(
            row,
            injectable_names=injectable_names,
            observable_names=observable_names,
        )
        if not roles:
            continue
        records.append(
            _canonical_record(
                row,
                platform=platform_name,
                roles=roles,
            )
        )
        seen.add(canonical_name)
    _validate_unique_record_identifiers(records)
    _validate_manifest_field_alignment(manifest, records)
    return sorted(records, key=lambda item: str(item["canonical_name"]))


def add_variable_aliases_to_injection_registry(
    registry: Mapping[str, Mapping[str, Any]],
    variable_map: list[Mapping[str, Any]],
) -> dict[str, dict[str, Any]]:
    enriched = {str(name): dict(spec) for name, spec in registry.items()}
    owners = {name: name for name in enriched}
    for record in variable_map:
        roles = {str(role).strip() for role in record.get("roles", [])}
        if "injectable" not in roles:
            continue
        manifest_name = str(record.get("manifest_name", "")).strip()
        if not manifest_name:
            raise ValueError(f"injectable variable missing manifest_name: {record}")
        if manifest_name not in enriched:
            raise ValueError(f"injectable variable missing runtime registry point: {manifest_name}")
        spec = enriched[manifest_name]
        for name in _record_names(record):
            owner = owners.get(name)
            if owner is not None and owner != manifest_name:
                raise ValueError(
                    f"ambiguous injection variable identifier {name}: {owner}, {manifest_name}"
                )
            enriched[name] = spec
            owners[name] = manifest_name
    return enriched


def build_te_variable_map(root: Path) -> list[dict[str, Any]]:
    variable_map = build_platform_variable_map(root, "tennessee_eastman")
    return [record for record in variable_map if record["layer"] == "setpoint"]


def resolve_variable(
    variable_map: list[Mapping[str, Any]],
    name: str,
) -> dict[str, Any]:
    target = str(name).strip()
    for record in variable_map:
        if target in _record_names(record):
            return dict(record)
        aliases = [str(alias).strip() for alias in record.get("aliases", [])]
        if target and target in aliases:
            return dict(record)
    raise KeyError(target)
